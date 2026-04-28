"""Excel export of entry/exit logs and barrier actions, with filters."""
from __future__ import annotations

import io
from datetime import datetime

from fastapi import APIRouter, Depends, Query, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.database import get_db
from src.models import BarrierActionLog, EntryExitLog
from src.routers.logs import _build_entry_query
from src.security import require_admin
from src.utils.query import OptInt, OptStr
from src.utils.tz import (
    local_day_range,
    local_month_range,
    local_year_range,
    now_local,
    to_local,
)

router = APIRouter(prefix="/api/export", tags=["export"])

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_HEADER_FILL = PatternFill("solid", fgColor="1E222B")
_HEADER_FONT = Font(bold=True, color="EEF0F3")

# Uzbek translations for enum values that end up in the spreadsheet.
_DIRECTION_UZ = {"entry": "Kirish", "exit": "Chiqish"}
_ACTION_UZ = {
    "open": "Ochildi",
    "close": "Yopildi",
    "lock": "Qulflandi",
    "unlock": "Qulfdan ochildi",
}
_ACTOR_UZ = {
    "admin": "Admin",
    "telegram": "Telegram",
    "auto_anpr": "ANPR (avto)",
    "auto_loop": "Avto-yopish",
}


def _enum_value(v) -> str:
    return v.value if hasattr(v, "value") else str(v) if v is not None else ""


def _translate(mapping: dict[str, str], v) -> str:
    raw = _enum_value(v)
    return mapping.get(raw, raw)


def _style_header_row(ws, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT


def _filename(prefix: str, plate: str | None, year, month, day) -> str:
    parts = [prefix]
    if year:
        parts.append(str(year))
    if month:
        parts.append(f"{int(month):02d}")
    if day:
        parts.append(f"{int(day):02d}")
    if plate:
        parts.append(plate)
    parts.append(now_local().strftime("%H%M%S"))
    return "_".join(parts) + ".xlsx"


@router.get("/entries.xlsx")
async def export_entries_xlsx(
    request: Request,
    plate: OptStr = None,
    direction: OptStr = None,
    camera: OptStr = None,
    date: OptStr = None,
    year: OptInt = None,
    month: OptInt = None,
    day: OptInt = None,
    start: datetime | None = Query(default=None),
    end: datetime | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    require_admin(request)
    if date:
        from src.utils.query import parse_iso_date

        year, month, day = parse_iso_date(date)
    stmt = _build_entry_query(
        plate, direction, year, month, day, start, end, camera
    )
    rows = list((await db.scalars(stmt)).all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Kirish-Chiqish"
    headers = [
        "ID",
        "Mashina raqami",
        "Yo‘nalish",
        "Kamera",
        "Kamera IP",
        "Aniqlik %",
        "Ruxsat",
        "Raqam rangi",
        "Vaqt (GMT+5)",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for r in rows:
        ws.append(
            [
                r.id,
                r.plate_number,
                _translate(_DIRECTION_UZ, r.direction),
                r.camera_name,
                r.camera_host,
                r.confidence,
                "Ha" if r.is_allowed else "Yo‘q",
                r.plate_color,
                to_local(r.event_time).replace(tzinfo=None) if r.event_time else None,
            ]
        )

    for col_idx, width in enumerate([8, 18, 12, 14, 16, 11, 10, 14, 22], start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type=_XLSX_MIME,
        headers={
            "Content-Disposition": (
                f'attachment; filename="{_filename("entries", plate, year, month, day)}"'
            )
        },
    )


@router.get("/barrier.xlsx")
async def export_barrier_xlsx(
    request: Request,
    actor: OptStr = None,
    action: OptStr = None,
    camera: OptStr = None,
    date: OptStr = None,
    year: OptInt = None,
    month: OptInt = None,
    day: OptInt = None,
    start: datetime | None = Query(default=None),
    end: datetime | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """Export barrier_action_logs filtered by actor/action/camera/date range."""
    require_admin(request)

    stmt = select(BarrierActionLog).order_by(BarrierActionLog.created_at.desc())
    # `date` (YYYY-MM-DD) overrides explicit year/month/day for convenience.
    if date:
        from src.utils.query import parse_iso_date

        year, month, day = parse_iso_date(date)
    if actor:
        stmt = stmt.where(BarrierActionLog.actor == actor)
    if action:
        stmt = stmt.where(BarrierActionLog.action == action)
    if camera:
        stmt = stmt.where(BarrierActionLog.camera_name.ilike(f"%{camera}%"))
    if year and month and day:
        s, e = local_day_range(year, month, day)
        stmt = stmt.where(
            BarrierActionLog.created_at >= s,
            BarrierActionLog.created_at < e,
        )
    elif year and month:
        s, e = local_month_range(year, month)
        stmt = stmt.where(
            BarrierActionLog.created_at >= s,
            BarrierActionLog.created_at < e,
        )
    elif year:
        s, e = local_year_range(year)
        stmt = stmt.where(
            BarrierActionLog.created_at >= s,
            BarrierActionLog.created_at < e,
        )
    if start:
        stmt = stmt.where(BarrierActionLog.created_at >= start)
    if end:
        stmt = stmt.where(BarrierActionLog.created_at <= end)

    rows = list((await db.scalars(stmt)).all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Shlakbaum"
    headers = [
        "ID",
        "Kamera",
        "Amal",
        "Kim",
        "Tafsilot",
        "Mashina raqami",
        "Status kodi",
        "Status xabari",
        "Vaqt (GMT+5)",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for r in rows:
        ws.append(
            [
                r.id,
                r.camera_name,
                _translate(_ACTION_UZ, r.action),
                _translate(_ACTOR_UZ, r.actor),
                r.actor_detail,
                r.plate_number,
                r.status_code,
                r.status_message,
                to_local(r.created_at).replace(tzinfo=None) if r.created_at else None,
            ]
        )

    for col_idx, width in enumerate(
        [6, 14, 14, 14, 28, 16, 12, 28, 22], start=1
    ):
        ws.column_dimensions[chr(64 + col_idx)].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type=_XLSX_MIME,
        headers={
            "Content-Disposition": (
                f'attachment; filename="{_filename("barrier", None, year, month, day)}"'
            )
        },
    )
