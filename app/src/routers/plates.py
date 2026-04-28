from __future__ import annotations

import io
import logging
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from src.database import get_db
from src.models import Plate
from src.schemas import PlateCreate, PlateOut, PlateUpdate
from src.security import require_admin
from src.services import plate_sync
from src.utils.tz import now_local, to_local

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/plates", tags=["plates"])

# Header layout used by both the downloadable template and the importer.
_TEMPLATE_HEADERS = ["plate_number", "owner_name", "is_allowed", "note"]
_TEMPLATE_EXAMPLES = [
    ["01A123BC", "Ali Valiyev", "yes", "PDP xodimi"],
    ["10B777CC", "Vali Aliyev", "no", "Bloklangan"],
]
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _truthy(val: Any) -> bool:
    """Lenient yes/no parser used while importing the Excel file."""
    if isinstance(val, bool):
        return val
    if val is None:
        return True
    s = str(val).strip().lower()
    return s in {"1", "true", "t", "yes", "y", "ha", "ruxsat", "allow", "allowed"}


@router.get("", response_model=list[PlateOut])
async def list_plates(
    request: Request,
    q: str | None = None,
    db: AsyncSession = Depends(get_db),
) -> list[Plate]:
    require_admin(request)
    stmt = select(Plate).order_by(Plate.created_at.desc())
    if q:
        stmt = stmt.where(Plate.plate_number.ilike(f"%{q.upper()}%"))
    return list((await db.scalars(stmt)).all())


@router.post("", response_model=PlateOut, status_code=201)
async def create_plate(
    payload: PlateCreate,
    request: Request,
    db: AsyncSession = Depends(get_db),
) -> Plate:
    require_admin(request)
    plate = Plate(
        plate_number=payload.plate_number.upper().strip(),
        owner_name=payload.owner_name,
        is_allowed=payload.is_allowed,
        note=payload.note,
    )
    db.add(plate)
    try:
        await db.commit()
    except IntegrityError:
        await db.rollback()
        raise HTTPException(status_code=409, detail="plate already exists")
    await db.refresh(plate)
    await plate_sync.push_plate(plate)
    return plate


@router.post("/sync-all")
async def sync_all_plates(
    request: Request, db: AsyncSession = Depends(get_db)
) -> dict:
    """Push every plate in our DB to both cameras. Useful after plates were
    added before the sync mechanism existed, or after a camera factory reset.
    """
    require_admin(request)
    rows = list((await db.scalars(select(Plate))).all())
    results = await plate_sync.push_plates_bulk(rows)
    return {
        "total": len(rows),
        "results": [
            {"camera": name, "ok": ok, "message": msg} for name, ok, msg in results
        ],
    }


# ---------- Excel import / template ----------
@router.get("/template.xlsx")
async def download_template(request: Request) -> StreamingResponse:
    """Download a blank xlsx template that the importer accepts as-is."""
    require_admin(request)
    wb = Workbook()
    ws = wb.active
    ws.title = "plates"
    ws.append(_TEMPLATE_HEADERS)
    for row in _TEMPLATE_EXAMPLES:
        ws.append(row)

    notes = wb.create_sheet("instructions")
    notes.append(["Field", "Description"])
    notes.append(["plate_number", "Talab. Lotin harfda, masalan 01A123BC."])
    notes.append([
        "owner_name",
        "Ixtiyoriy. Egasining ismi va familiyasi.",
    ])
    notes.append([
        "is_allowed",
        "yes/no, true/false, 1/0. Bo'sh — yes (ruxsat).",
    ])
    notes.append(["note", "Ixtiyoriy izoh."])
    for col_idx, width in enumerate([16, 24, 12, 30], start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width
    notes.column_dimensions["A"].width = 16
    notes.column_dimensions["B"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type=_XLSX_MIME,
        headers={
            "Content-Disposition": 'attachment; filename="plates_template.xlsx"'
        },
    )


@router.post("/import")
async def import_plates(
    request: Request,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Import plates from an Excel (.xlsx) file.

    On every imported / updated row we ALSO push the change to all cameras
    via ISAPI §11.6.5 so the device-side allowlist stays in sync.
    """
    require_admin(request)

    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="xlsx file required")

    raw = await file.read()
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"invalid xlsx: {exc}") from exc

    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(v).strip().lower() if v is not None else "" for v in next(rows)]
    except StopIteration:
        raise HTTPException(status_code=400, detail="empty file")

    # Map column name → index (allow extra/missing optional columns).
    col = {name: header.index(name) for name in _TEMPLATE_HEADERS if name in header}
    if "plate_number" not in col:
        raise HTTPException(
            status_code=400,
            detail="header must contain 'plate_number' column",
        )

    created: list[Plate] = []
    updated: list[Plate] = []
    skipped: list[dict[str, Any]] = []

    seen: set[str] = set()
    for line_no, row in enumerate(rows, start=2):
        if not row:
            continue
        raw_number = row[col["plate_number"]]
        if raw_number is None or str(raw_number).strip() == "":
            continue
        plate_number = str(raw_number).strip().upper()
        if len(plate_number) < 2 or len(plate_number) > 32:
            skipped.append(
                {"line": line_no, "plate": plate_number, "reason": "bad length"}
            )
            continue
        if plate_number in seen:
            skipped.append(
                {"line": line_no, "plate": plate_number, "reason": "duplicate row"}
            )
            continue
        seen.add(plate_number)

        owner = row[col["owner_name"]] if "owner_name" in col else None
        is_allowed = (
            _truthy(row[col["is_allowed"]]) if "is_allowed" in col else True
        )
        note = row[col["note"]] if "note" in col else None

        existing = (
            await db.scalars(
                select(Plate).where(Plate.plate_number == plate_number)
            )
        ).first()

        if existing is None:
            plate = Plate(
                plate_number=plate_number,
                owner_name=str(owner).strip() if owner else None,
                is_allowed=is_allowed,
                note=str(note).strip() if note else None,
            )
            db.add(plate)
            try:
                await db.flush()
            except IntegrityError:
                await db.rollback()
                skipped.append(
                    {
                        "line": line_no,
                        "plate": plate_number,
                        "reason": "integrity error",
                    }
                )
                continue
            created.append(plate)
        else:
            if owner is not None:
                existing.owner_name = str(owner).strip() or None
            if "is_allowed" in col:
                existing.is_allowed = is_allowed
            if note is not None:
                existing.note = str(note).strip() or None
            updated.append(existing)

    await db.commit()
    for p in created + updated:
        await db.refresh(p)

    # Push everything we just touched to both cameras in a single batched call.
    push_results = await plate_sync.push_plates_bulk(created + updated)

    return {
        "created": len(created),
        "updated": len(updated),
        "skipped": skipped,
        "camera_push": [
            {"camera": name, "ok": ok, "message": msg}
            for name, ok, msg in push_results
        ],
    }


# ---------- Excel export ----------
@router.get("/export.xlsx")
async def export_plates(
    request: Request,
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """Export the entire plates table as xlsx."""
    require_admin(request)
    rows = list(
        (await db.scalars(select(Plate).order_by(Plate.created_at.desc()))).all()
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Raqamlar"
    ws.append(
        [
            "ID",
            "Mashina raqami",
            "Egasi",
            "Ruxsat",
            "Izoh",
            "Qo‘shilgan (GMT+5)",
        ]
    )
    for r in rows:
        ws.append(
            [
                r.id,
                r.plate_number,
                r.owner_name,
                "Ha" if r.is_allowed else "Yo‘q",
                r.note,
                to_local(r.created_at).replace(tzinfo=None) if r.created_at else None,
            ]
        )
    for col_idx, width in enumerate([6, 18, 24, 10, 30, 22], start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"plates_{now_local().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type=_XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------- Single-plate CRUD ----------
@router.get("/{plate_id}", response_model=PlateOut)
async def get_plate(
    plate_id: int, request: Request, db: AsyncSession = Depends(get_db)
) -> Plate:
    require_admin(request)
    plate = await db.get(Plate, plate_id)
    if not plate:
        raise HTTPException(status_code=404, detail="not found")
    return plate


@router.put("/{plate_id}", response_model=PlateOut)
async def update_plate(
    plate_id: int,
    payload: PlateUpdate,
    request: Request,
    db: AsyncSession = Depends(get_db),
) -> Plate:
    require_admin(request)
    plate = await db.get(Plate, plate_id)
    if not plate:
        raise HTTPException(status_code=404, detail="not found")
    data = payload.model_dump(exclude_unset=True)
    if "plate_number" in data and data["plate_number"]:
        data["plate_number"] = data["plate_number"].upper().strip()
    for k, v in data.items():
        setattr(plate, k, v)
    try:
        await db.commit()
    except IntegrityError:
        await db.rollback()
        raise HTTPException(status_code=409, detail="duplicate plate number")
    await db.refresh(plate)
    await plate_sync.push_plate(plate)
    return plate


@router.delete("/{plate_id}", status_code=204)
async def delete_plate(
    plate_id: int, request: Request, db: AsyncSession = Depends(get_db)
):
    require_admin(request)
    plate = await db.get(Plate, plate_id)
    if not plate:
        raise HTTPException(status_code=404, detail="not found")
    plate_number = plate.plate_number
    await db.delete(plate)
    await db.commit()
    await plate_sync.delete_plate(plate_number)
