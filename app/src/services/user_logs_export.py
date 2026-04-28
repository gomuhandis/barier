"""Build a per-user Excel report of entry/exit events and barrier actions.

Used by the Telegram bot when a user requests to download their own logs.
The same Uzbek translations as the admin /api/export/* endpoints are used so
both reports look identical.
"""
from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.models import (
    BarrierAction,
    BarrierActionLog,
    BarrierActor,
    EntryExitLog,
    TelegramUser,
)
from src.utils.tz import to_local

_HEADER_FILL = PatternFill("solid", fgColor="1E222B")
_HEADER_FONT = Font(bold=True, color="EEF0F3")

_DIRECTION_UZ = {"entry": "Kirish", "exit": "Chiqish"}
_ACTION_UZ = {
    "open": "Ochildi",
    "close": "Yopildi",
    "lock": "Qulflandi",
    "unlock": "Qulfdan ochildi",
}


def _enum_value(v) -> str:
    if v is None:
        return ""
    return v.value if hasattr(v, "value") else str(v)


def _style_header(ws, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT


async def build_user_logs_xlsx(
    session: AsyncSession,
    *,
    user: TelegramUser,
    plate_numbers: list[str],
    start: datetime,
    end: datetime,
    label: str,
) -> bytes:
    """Return an .xlsx file as bytes containing two sheets:

    1. *Kirish-Chiqish* — every ANPR event for the user's linked cars in
       the [start, end) window.
    2. *Mening tugmalarim* — every barrier action this user triggered via
       the bot in the same window (matched by phone in actor_detail).
    """
    if plate_numbers:
        entries = list(
            (
                await session.scalars(
                    select(EntryExitLog)
                    .where(
                        EntryExitLog.plate_number.in_(plate_numbers),
                        EntryExitLog.event_time >= start,
                        EntryExitLog.event_time < end,
                    )
                    .order_by(EntryExitLog.event_time.desc())
                )
            ).all()
        )
    else:
        entries = []

    actions = list(
        (
            await session.scalars(
                select(BarrierActionLog)
                .where(
                    BarrierActionLog.actor == BarrierActor.TELEGRAM,
                    BarrierActionLog.actor_detail.ilike(
                        f"%phone={user.phone_number}%"
                    ),
                    BarrierActionLog.created_at >= start,
                    BarrierActionLog.created_at < end,
                )
                .order_by(BarrierActionLog.created_at.desc())
            )
        ).all()
    )

    wb = Workbook()

    # Sheet 1: car entries / exits
    ws1 = wb.active
    ws1.title = "Kirish-Chiqish"
    ws1.append(
        [
            "Vaqt (GMT+5)",
            "Mashina raqami",
            "Yo‘nalish",
            "Kamera",
            "Aniqlik %",
            "Ruxsat",
        ]
    )
    _style_header(ws1, 6)
    for r in entries:
        direction = _DIRECTION_UZ.get(_enum_value(r.direction), _enum_value(r.direction))
        ws1.append(
            [
                to_local(r.event_time).replace(tzinfo=None) if r.event_time else None,
                r.plate_number,
                direction,
                r.camera_name,
                r.confidence,
                "Ha" if r.is_allowed else "Yo‘q",
            ]
        )
    for col_idx, width in enumerate([22, 18, 12, 14, 11, 10], start=1):
        ws1.column_dimensions[chr(64 + col_idx)].width = width
    ws1.freeze_panes = "A2"

    # Sheet 2: this user's barrier actions
    ws2 = wb.create_sheet("Mening tugmalarim")
    ws2.append(
        [
            "Vaqt (GMT+5)",
            "Kamera",
            "Amal",
            "Status kodi",
            "Status xabari",
        ]
    )
    _style_header(ws2, 5)
    for r in actions:
        action = _ACTION_UZ.get(_enum_value(r.action), _enum_value(r.action))
        ws2.append(
            [
                to_local(r.created_at).replace(tzinfo=None) if r.created_at else None,
                r.camera_name,
                action,
                r.status_code,
                r.status_message,
            ]
        )
    for col_idx, width in enumerate([22, 14, 14, 12, 30], start=1):
        ws2.column_dimensions[chr(64 + col_idx)].width = width
    ws2.freeze_panes = "A2"

    # Sheet 3: summary
    ws3 = wb.create_sheet("Hisobot")
    ws3.append(["Maydon", "Qiymat"])
    _style_header(ws3, 2)
    ws3.append(["Foydalanuvchi", user.full_name or user.phone_number])
    ws3.append(["Telefon", user.phone_number])
    ws3.append(["Mashinalar", ", ".join(plate_numbers) if plate_numbers else "—"])
    ws3.append(["Davr", label])
    ws3.append(
        [
            "Boshlanish",
            to_local(start).strftime("%Y-%m-%d %H:%M") if start else "—",
        ]
    )
    ws3.append(
        ["Tugash", to_local(end).strftime("%Y-%m-%d %H:%M") if end else "—"]
    )
    ws3.append(["Kirish/Chiqish soni", len(entries)])
    entries_in = sum(1 for r in entries if _enum_value(r.direction) == "entry")
    entries_out = sum(1 for r in entries if _enum_value(r.direction) == "exit")
    ws3.append(["  — Kirgan", entries_in])
    ws3.append(["  — Chiqgan", entries_out])
    ws3.append(["Shlakbaum tugmalari", len(actions)])
    opens = sum(1 for r in actions if r.action == BarrierAction.OPEN)
    closes = sum(1 for r in actions if r.action == BarrierAction.CLOSE)
    ws3.append(["  — Ochishlar", opens])
    ws3.append(["  — Yopishlar", closes])
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 36

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
